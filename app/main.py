from fastapi import FastAPI, File, HTTPException, UploadFile, Request
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from io import BytesIO
import statistics
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageStat

# Optional OCR dependency
try:
    import pytesseract
    from pytesseract import Output as TesseractOutput, get_tesseract_version

    _HAS_PYTESSERACT = True
    try:
        get_tesseract_version()
        _HAS_TESSERACT = True
    except Exception:
        _HAS_TESSERACT = False
except Exception:
    _HAS_PYTESSERACT = False
    _HAS_TESSERACT = False

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = BASE_DIR / "templates"

app = FastAPI(title="PDF Table Extractor")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

ALLOWED_CONTENT_TYPES = {
    "application/pdf",
    "application/x-pdf",
    "application/octet-stream",  # some browsers send this for PDF uploads
}

DEFAULT_TABLE_SETTINGS: Dict[str, object] = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
    "snap_tolerance": 3,
    "join_tolerance": 3,
}


def _ensure_pdf_bytes(pdf_bytes: bytes) -> None:
    if not pdf_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")


def _sanitize_filename(filename: Optional[str]) -> str:
    if not filename:
        return "tables"
    stem = Path(filename).stem
    safe = "".join(ch for ch in stem if ch.isalnum() or ch in {"-", "_"})
    return (safe or "tables")[:60]


def _color_to_hex(color: Optional[List[float]]) -> Optional[str]:
    """Convert a PDF RGB/gray color array to 8-char ARGB hex for openpyxl."""
    if not color:
        return None
    try:
        if len(color) == 1:
            r = g = b = int(color[0] * 255)
        else:
            r, g, b = (int(c * 255) for c in color[:3])
        return f"FF{r:02X}{g:02X}{b:02X}"
    except Exception:
        return None


def _rectangles_with_color(page) -> List[Dict[str, object]]:
    """Collect rectangles/highlight annotations that carry a fill/stroke color."""
    rects = []
    for rect in getattr(page, "rects", []):
        if rect.get("non_stroking_color") or rect.get("stroke_color"):
            rects.append(rect)

    for annot in getattr(page, "annots", []) or []:
        if annot.get("subtype", "").lower() == "highlight":
            rects.append(
                {
                    "x0": annot.get("x0"),
                    "x1": annot.get("x1"),
                    "top": annot.get("top") or annot.get("y1"),
                    "bottom": annot.get("bottom") or annot.get("y0"),
                    "non_stroking_color": annot.get("color"),
                }
            )
    return rects


def _normalize_box(obj: Dict[str, object]) -> Optional[Dict[str, float]]:
    """Return a consistent box dict or None if coordinates are missing."""
    x0 = obj.get("x0")
    x1 = obj.get("x1")
    top = obj.get("top") if obj.get("top") is not None else obj.get("y1")
    bottom = obj.get("bottom") if obj.get("bottom") is not None else obj.get("y0")
    if None in (x0, x1, top, bottom):
        return None
    return {"x0": float(x0), "x1": float(x1), "top": float(top), "bottom": float(bottom)}


def _boxes_overlap(a: Dict[str, float], b: Dict[str, float]) -> bool:
    """Basic AABB overlap check."""
    return not (
        a["x1"] <= b["x0"]
        or a["x0"] >= b["x1"]
        or a["top"] >= b["bottom"]
        or a["bottom"] <= b["top"]
    )


def _map_highlights(table, page) -> List[List[Optional[str]]]:
    """Return a matrix of ARGB colors aligned with extracted table rows.

    pdfplumber 0.11 switched `Table.cols` to `Table.columns` and now returns
    `Row`/`Column` objects with bounding boxes. This helper tolerates both the
    older numeric edges API and the newer object-based API.
    """

    data = table.extract()
    if not data:
        return []

    rows = len(data)
    cols = max(len(r) for r in data)
    highlights: List[List[Optional[str]]] = [
        [None for _ in range(cols)] for _ in range(rows)
    ]

    rects = _rectangles_with_color(page)
    if not rects:
        return highlights

    table_rows = getattr(table, "rows", []) or []
    table_cols = (
        getattr(table, "cols", None)
        if hasattr(table, "cols")
        else getattr(table, "columns", [])
    ) or []

    def _row_bounds(idx: int) -> Optional[Tuple[float, float]]:
        try:
            row_obj = table_rows[idx]
        except Exception:
            return None

        # Newer pdfplumber returns Row objects with bbox
        if hasattr(row_obj, "bbox"):
            x0, top, x1, bottom = row_obj.bbox
            return float(top), float(bottom)

        # Fallback: list/tuple shaped like a bbox
        if isinstance(row_obj, (list, tuple)) and len(row_obj) >= 4:
            return float(row_obj[1]), float(row_obj[3])

        # Fallback: numeric edge list (edges length = rows + 1)
        if isinstance(row_obj, (int, float)):
            try:
                nxt = table_rows[idx + 1]
                if isinstance(nxt, (int, float)):
                    top, bottom = float(row_obj), float(nxt)
                    return (top, bottom) if top <= bottom else (bottom, top)
            except Exception:
                return None
        return None

    def _col_bounds(idx: int) -> Optional[Tuple[float, float]]:
        try:
            col_obj = table_cols[idx]
        except Exception:
            return None

        if hasattr(col_obj, "bbox"):
            x0, top, x1, bottom = col_obj.bbox
            return float(x0), float(x1)

        if isinstance(col_obj, (list, tuple)) and len(col_obj) >= 4:
            return float(col_obj[0]), float(col_obj[2])

        if isinstance(col_obj, (int, float)):
            try:
                nxt = table_cols[idx + 1]
                if isinstance(nxt, (int, float)):
                    x0, x1 = float(col_obj), float(nxt)
                    return (x0, x1) if x0 <= x1 else (x1, x0)
            except Exception:
                return None
        return None

    for r_idx, row_cells in enumerate(data):
        bounds = _row_bounds(r_idx)
        if not bounds:
            continue
        row_top, row_bottom = bounds

        for c_idx, _ in enumerate(row_cells):
            cbounds = _col_bounds(c_idx)
            if not cbounds:
                continue
            x0, x1 = cbounds
            cell_box = {"x0": x0, "x1": x1, "top": row_top, "bottom": row_bottom}

            for rect in rects:
                rcolor = _color_to_hex(
                    rect.get("non_stroking_color") or rect.get("stroke_color")
                )
                rect_box = _normalize_box(rect)
                if rcolor and rect_box and _boxes_overlap(cell_box, rect_box):
                    highlights[r_idx][c_idx] = rcolor
                    break

    return highlights


def _page_has_text(page) -> bool:
    """Quickly determine if the PDF page already contains extractable text."""
    try:
        return bool(getattr(page, "chars", []) or getattr(page, "objects", {}).get("chars"))
    except Exception:
        return False


def _to_page_image(page, resolution: int = 400) -> Optional[Image.Image]:
    """Render a PDF page to a PIL image."""
    try:
        return page.to_image(resolution=resolution).original.convert("RGB")
    except Exception:
        return None


def _pad_bbox(box: Dict[str, float], pad: int, max_width: int, max_height: int) -> Dict[str, int]:
    """Pad a bounding box while keeping it inside the image."""
    return {
        "x0": max(0, int(box["x0"] - pad)),
        "top": max(0, int(box["top"] - pad)),
        "x1": min(max_width, int(box["x1"] + pad)),
        "bottom": min(max_height, int(box["bottom"] + pad)),
    }


def _estimate_highlight_color(image: Image.Image, bbox: Dict[str, float]) -> Optional[str]:
    """Estimate a highlight color by averaging pixels inside the bbox."""
    if image is None:
        return None

    width, height = image.size
    padded = _pad_bbox(bbox, pad=2, max_width=width, max_height=height)
    region = image.crop((padded["x0"], padded["top"], padded["x1"], padded["bottom"]))
    if region.size[0] == 0 or region.size[1] == 0:
        return None

    stat = ImageStat.Stat(region)
    r, g, b = stat.mean[:3]
    brightness = (r + g + b) / 3.0
    maxc = max(r, g, b)
    minc = min(r, g, b)
    saturation = 0 if maxc == 0 else (maxc - minc) / maxc

    # Treat moderately bright + colorful regions as highlights; ignore grayscale/white.
    if saturation > 0.25 and brightness > 120:
        return f"FF{int(r):02X}{int(g):02X}{int(b):02X}"
    return None


def _ocr_rows_and_highlights(image: Image.Image) -> Tuple[List[List[str]], List[List[Optional[str]]]]:
    """Use Tesseract to recover rows/columns from a rendered page image."""
    if not (_HAS_PYTESSERACT and _HAS_TESSERACT) or image is None:
        return [], []

    try:
        data = pytesseract.image_to_data(
            image, output_type=TesseractOutput.DICT, config="--psm 3"
        )
    except Exception:
        return [], []

    words = []
    n = len(data.get("text", []))
    for idx in range(n):
        text = (data["text"][idx] or "").strip()
        conf_raw = data.get("conf", ["-1"])[idx]
        try:
            conf = int(float(conf_raw))
        except Exception:
            conf = -1
        if not text or conf < 40:  # skip low-confidence noise
            continue

        x0 = int(data["left"][idx])
        y0 = int(data["top"][idx])
        w = int(data["width"][idx])
        h = int(data["height"][idx])

        words.append(
            {
                "text": text,
                "x0": x0,
                "x1": x0 + w,
                "top": y0,
                "bottom": y0 + h,
                "block": data.get("block_num", [0])[idx],
                "par": data.get("par_num", [0])[idx],
                "line": data.get("line_num", [0])[idx],
            }
        )

    if not words:
        return [], []

    rows: List[List[str]] = []
    highlights: List[List[Optional[str]]] = []

    # Group words into rows by y-position tolerance.
    median_height = statistics.median([w["bottom"] - w["top"] for w in words])
    row_threshold = max(25, median_height * 0.8)
    words_sorted = sorted(words, key=lambda w: w["top"])

    current_row: List[Dict[str, object]] = []
    current_top: Optional[float] = None

    def _flush_row(row_words: List[Dict[str, object]]):
        if not row_words:
            return
        row_words.sort(key=lambda w: w["x0"])
        gaps = [
            max(0, b["x0"] - a["x1"]) for a, b in zip(row_words, row_words[1:])
        ]
        median_gap = statistics.median(gaps) if gaps else 25
        gap_threshold = max(18, min(120, median_gap * 0.8))

        row_cells: List[str] = []
        row_colors: List[Optional[str]] = []
        cell_words: List[Dict[str, object]] = []
        prev_x1: Optional[int] = None

        for word in row_words:
            if prev_x1 is not None and (word["x0"] - prev_x1) > gap_threshold:
                cell_text = " ".join(w["text"] for w in cell_words).strip()
                bbox = {
                    "x0": min(w["x0"] for w in cell_words),
                    "x1": max(w["x1"] for w in cell_words),
                    "top": min(w["top"] for w in cell_words),
                    "bottom": max(w["bottom"] for w in cell_words),
                }
                row_cells.append(cell_text)
                row_colors.append(_estimate_highlight_color(image, bbox))
                cell_words = []

            cell_words.append(word)
            prev_x1 = word["x1"]

        if cell_words:
            cell_text = " ".join(w["text"] for w in cell_words).strip()
            bbox = {
                "x0": min(w["x0"] for w in cell_words),
                "x1": max(w["x1"] for w in cell_words),
                "top": min(w["top"] for w in cell_words),
                "bottom": max(w["bottom"] for w in cell_words),
            }
            row_cells.append(cell_text)
            row_colors.append(_estimate_highlight_color(image, bbox))

        rows.append(row_cells)
        highlights.append(row_colors)

    for word in words_sorted:
        if current_top is None:
            current_top = word["top"]
        if word["top"] - current_top > row_threshold:
            _flush_row(current_row)
            current_row = [word]
            current_top = word["top"]
        else:
            current_row.append(word)
    _flush_row(current_row)

    # Normalize column counts so downstream code can rely on rectangular data.
    max_cols = max(len(r) for r in rows) if rows else 0
    for idx, row in enumerate(rows):
        pad_len = max_cols - len(row)
        if pad_len > 0:
            row.extend([""] * pad_len)
            highlights[idx].extend([None] * pad_len)

    return rows, highlights


def _extract_tables_via_ocr(page, page_idx: int) -> List[Dict[str, object]]:
    """Fallback: OCR a rendered page when no vector tables/text are present."""
    if not (_HAS_PYTESSERACT and _HAS_TESSERACT):
        return []

    image = _to_page_image(page)
    if image is None:
        return []

    rows, highlights = _ocr_rows_and_highlights(image)
    if not rows:
        return []

    return [
        {
            "title": f"page-{page_idx}-ocr-1",
            "rows": rows,
            "highlights": highlights,
        }
    ]


def parse_pdf_tables(
    pdf_bytes: bytes, table_settings: Optional[Dict[str, object]] = None
) -> List[Dict[str, object]]:
    """Parse tables with optional highlight colors."""
    settings = table_settings or DEFAULT_TABLE_SETTINGS
    results: List[Dict[str, object]] = []
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            tables = page.find_tables(table_settings=settings)

            # OCR fallback for image-only or table-less pages.
            if not tables and not _page_has_text(page):
                ocr_tables = _extract_tables_via_ocr(page, page_idx)
                if ocr_tables:
                    results.extend(ocr_tables)
                    continue

            for table_idx, table in enumerate(tables, start=1):
                data = table.extract()
                if not data:
                    continue
                highlights = _map_highlights(table, page)
                results.append(
                    {
                        "title": f"page-{page_idx}-table-{table_idx}",
                        "rows": data,
                        "highlights": highlights,
                    }
                )
    return results


def extract_tables_to_workbook(
    pdf_bytes: bytes, table_settings: Optional[Dict[str, object]] = None
) -> Tuple[BytesIO, int]:
    """Return an Excel workbook stream plus detected table count."""

    try:
        tables = parse_pdf_tables(pdf_bytes, table_settings)
        workbook = Workbook()
        sheet_count = 0

        for table in tables:
            worksheet = workbook.active if sheet_count == 0 else workbook.create_sheet()
            worksheet.title = table["title"]
            rows = table["rows"]
            highlights = table.get("highlights") or []
            for r_idx, row in enumerate(rows):
                worksheet.append(row)
                for c_idx, _ in enumerate(row):
                    try:
                        color = highlights[r_idx][c_idx]
                    except Exception:
                        color = None
                    if color:
                        worksheet.cell(row=r_idx + 1, column=c_idx + 1).fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
            sheet_count += 1

        if sheet_count == 0:
            worksheet = workbook.active
            worksheet.title = "no-tables-found"
            worksheet.append(["No tables were detected in this PDF."])

        output_stream = BytesIO()
        workbook.save(output_stream)
        output_stream.seek(0)
        return output_stream, sheet_count
    except Exception as exc:  # pdf could be corrupted or unsupported
        raise HTTPException(status_code=400, detail=f"Failed to parse PDF: {exc}")


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


def _is_pdf(content_type: Optional[str]) -> bool:
    if not content_type:
        return False
    lowered = content_type.lower()
    return lowered.startswith("application/pdf") or lowered in ALLOWED_CONTENT_TYPES


@app.post("/analyze")
async def analyze_tables(file: UploadFile = File(...)):
    if not _is_pdf(file.content_type):
        raise HTTPException(status_code=400, detail="Please upload a PDF file.")

    pdf_bytes = await file.read()
    _ensure_pdf_bytes(pdf_bytes)
    tables = parse_pdf_tables(pdf_bytes, table_settings=DEFAULT_TABLE_SETTINGS)
    response = {
        "table_count": len(tables),
        # Send full rows/highlights for a complete on-page preview.
        "tables": [
            {
                "title": table["title"],
                "rows": table["rows"],
                "highlights": table.get("highlights") or [],
            }
            for table in tables
        ],
    }
    return JSONResponse(response)


@app.post("/extract")
async def extract_tables(file: UploadFile = File(...)):
    if not _is_pdf(file.content_type):
        raise HTTPException(status_code=400, detail="Please upload a PDF file.")

    pdf_bytes = await file.read()
    _ensure_pdf_bytes(pdf_bytes)

    excel_stream, table_count = extract_tables_to_workbook(
        pdf_bytes, table_settings=DEFAULT_TABLE_SETTINGS
    )
    filename = _sanitize_filename(file.filename)
    disposition = f"attachment; filename={filename}-tables.xlsx"

    headers = {
        "Content-Disposition": disposition,
        "X-Table-Count": str(table_count),
    }

    return StreamingResponse(
        content=excel_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)
