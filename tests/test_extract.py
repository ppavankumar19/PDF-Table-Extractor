from io import BytesIO

import pytest
from app.main import extract_tables_to_workbook, _HAS_TESSERACT
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont


def _build_sample_pdf() -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    data = [["Name", "Age"], ["Alice", "30"], ["Bob", "28"]]
    table = Table(data)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    doc.build([table])
    buffer.seek(0)
    return buffer.getvalue()


def test_extract_tables_to_workbook_finds_table():
    pdf_bytes = _build_sample_pdf()
    excel_stream, sheet_count = extract_tables_to_workbook(pdf_bytes)

    assert sheet_count == 1

    excel_stream.seek(0)
    workbook = load_workbook(filename=excel_stream)
    worksheet = workbook.active

    assert worksheet.title.startswith("page-1-table-1")

    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("Name", "Age")
    assert rows[1] == ("Alice", "30")
    assert rows[2] == ("Bob", "28")


def _build_image_table_pdf() -> bytes:
    """Create a simple 2x2 table as an embedded PNG to force OCR fallback."""
    img = Image.new("RGB", (1200, 800), "white")
    draw = ImageDraw.Draw(img)
    draw.rectangle((40, 40, 1160, 760), outline="black", width=8)
    draw.line((40, 400, 1160, 400), fill="black", width=6)
    draw.line((600, 40, 600, 760), fill="black", width=6)

    # Highlight bottom-left cell to test color pickup.
    draw.rectangle((48, 408, 592, 752), fill=(255, 255, 170))

    try:
        font = ImageFont.truetype("DejaVuSans.ttf", 80)
    except Exception:
        font = ImageFont.load_default()

    def _centered(text, cx, cy):
        bbox = draw.textbbox((0, 0), text, font=font)
        w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
        return cx - w // 2, cy - h // 2

    def _draw(text, cx, cy):
        draw.text(_centered(text, cx, cy), text, fill="black", font=font)

    _draw("Name", 320, 200)
    _draw("AGE", 880, 200)
    _draw("Alice", 320, 580)
    _draw("30", 880, 580)

    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    pdf_buffer = BytesIO()
    page_size = (1200, 800)
    c = canvas.Canvas(pdf_buffer, pagesize=page_size)
    c.drawImage(ImageReader(buf), 0, 0, width=1200, height=800, preserveAspectRatio=True)
    c.showPage()
    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()


@pytest.mark.skipif(not _HAS_TESSERACT, reason="Tesseract OCR is not available")
def test_extract_tables_from_image_pdf_via_ocr():
    pdf_bytes = _build_image_table_pdf()
    excel_stream, sheet_count = extract_tables_to_workbook(pdf_bytes)

    assert sheet_count == 1

    excel_stream.seek(0)
    workbook = load_workbook(filename=excel_stream)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    assert (rows[0][0] or "").lower() == "name"
    assert (rows[0][1] or "").lower() == "age"
    assert (rows[1][0] or "").lower() == "alice"
    assert (rows[1][1] or "").lower() == "30"
